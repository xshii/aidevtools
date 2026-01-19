#!/usr/bin/env python
"""xlsx → Python Transformer Demo

演示从 Excel 配置生成 Transformer 模型并运行:
1. 创建 xlsx 模板
2. 在 xlsx 中定义 Transformer 算子序列
3. 从 xlsx 生成 Python 代码
4. 运行并比对结果

使用方法:
    cd demos/05_xlsx_transformer
    python run.py
"""
import numpy as np
import sys
from pathlib import Path

# 添加 src 到 path
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "src"))


def create_transformer_xlsx(xlsx_path: str):
    """
    创建 Transformer 模型的 xlsx 配置

    模型结构 (简化版 1 层 Transformer):
        input_ids → embedding → linear(Q) → linear(K) → linear(V)
                              → attention → linear(O) → add(residual)
                              → layernorm → linear(FFN_up) → gelu
                              → linear(FFN_down) → add(residual) → layernorm
    """
    from aidevtools.xlsx import create_template

    # 创建模板，指定 transformer 相关算子
    create_template(xlsx_path, ops=[
        "embedding", "linear", "matmul", "attention",
        "softmax", "layernorm", "gelu", "add"
    ])

    # 编辑 xlsx 添加 transformer 算子配置
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws_ops = wb["ops"]

    # 清空示例数据
    ws_ops.delete_rows(3, ws_ops.max_row)

    # Transformer 配置 (简化版，1 层)
    # 格式: id, op_name, shape, dtype, depends, qtype, skip, sim_cmd, note
    # qtype: bfp4 用于 matmul/linear, bfp8 用于其他
    # sim_cmd: 留空（不执行仿真命令）
    transformer_ops = [
        # Embedding
        (0, "embedding", "1,16", "int32", "", "bfp8", "FALSE", "", "Token Embedding"),

        # Self-Attention: Q, K, V projections
        (1, "linear", "1,16,64", "float32", "0", "bfp4", "FALSE", "", "Q projection"),
        (2, "linear", "1,16,64", "float32", "0", "bfp4", "FALSE", "", "K projection"),
        (3, "linear", "1,16,64", "float32", "0", "bfp4", "FALSE", "", "V projection"),

        # Attention
        (4, "attention", "1,16,64", "float32", "1,2,3", "bfp8", "FALSE", "", "Attention"),

        # Output projection
        (5, "linear", "1,16,64", "float32", "4", "bfp4", "FALSE", "", "O projection"),

        # Residual + LayerNorm
        (6, "add", "1,16,64", "float32", "0,5", "bfp8", "FALSE", "", "Residual 1"),
        (7, "layernorm", "1,16,64", "float32", "6", "bfp8", "FALSE", "", "LayerNorm 1"),

        # FFN
        (8, "linear", "1,16,256", "float32", "7", "bfp4", "FALSE", "", "FFN up"),
        (9, "gelu", "1,16,256", "float32", "8", "bfp8", "FALSE", "", "GELU"),
        (10, "linear", "1,16,64", "float32", "9", "bfp4", "FALSE", "", "FFN down"),

        # Residual + LayerNorm
        (11, "add", "1,16,64", "float32", "7,10", "bfp8", "FALSE", "", "Residual 2"),
        (12, "layernorm", "1,16,64", "float32", "11", "bfp8", "FALSE", "", "Output LayerNorm"),
    ]

    for row_idx, config in enumerate(transformer_ops, 3):
        for col_idx, value in enumerate(config, 1):
            ws_ops.cell(row=row_idx, column=col_idx, value=value)

    wb.save(xlsx_path)
    print(f"    Transformer 配置已写入: {xlsx_path}")
    print(f"    共 {len(transformer_ops)} 个算子")

    return transformer_ops


def show_xlsx_content(xlsx_path: str):
    """显示 xlsx 内容"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)

    print("\n    [ops sheet - Transformer 算子配置]")
    print("    " + "-" * 80)

    ws_ops = wb["ops"]
    # 打印表头
    headers = [cell.value for cell in ws_ops[1]]
    print(f"    {headers}")
    print("    " + "-" * 80)

    # 打印数据
    for row in ws_ops.iter_rows(min_row=3, values_only=True):
        if row[0] is not None:  # 跳过空行
            # 格式化输出
            row_str = f"    [{row[0]:2}] {row[1]:12} shape={row[2]:12} qtype={row[5]:6} note={row[8]}"
            print(row_str)


def generate_python_code(xlsx_path: str, output_dir: str) -> str:
    """从 xlsx 生成 Python 代码"""
    from aidevtools.xlsx import import_xlsx

    py_path = Path(output_dir) / "generated_transformer.py"
    code = import_xlsx(xlsx_path, str(py_path))

    print(f"\n    生成 Python 代码: {py_path}")
    print("\n    [生成的代码预览]")
    print("    " + "-" * 60)

    lines = code.split("\n")
    for i, line in enumerate(lines[:30]):
        print(f"    {line}")
    if len(lines) > 30:
        print(f"    ... (共 {len(lines)} 行)")

    return str(py_path)


def run_transformer_xlsx(xlsx_path: str, output_dir: str):
    """运行 xlsx 配置的 Transformer"""
    from aidevtools.xlsx import run_xlsx

    print("\n    运行 Transformer 模型...")
    results = run_xlsx(xlsx_path, output_dir)

    print("\n    [运行结果]")
    print("    " + "-" * 60)
    print(f"    {'ID':>3} {'算子':12} {'状态':8} {'备注'}")
    print("    " + "-" * 60)

    pass_count = 0
    fail_count = 0
    skip_count = 0

    for r in results:
        status = r.get("status", "?")
        op_id = r.get("id", "?")
        op_name = r.get("op_name", "?")
        note = r.get("note", "")

        if status == "PASS":
            status_icon = "✓"
            pass_count += 1
        elif status == "FAIL":
            status_icon = "✗"
            fail_count += 1
        elif status == "SKIP":
            status_icon = "○"
            skip_count += 1
        else:
            status_icon = "?"

        print(f"    [{status_icon}] {op_id:2} {op_name:12} {status:8} {note}")

    print("    " + "-" * 60)
    print(f"    总计: {pass_count} PASS, {fail_count} FAIL, {skip_count} SKIP")

    return results


def main():
    """主函数"""
    print("""
╔══════════════════════════════════════════════════════════════════════╗
║              xlsx → Python Transformer Demo                          ║
╠══════════════════════════════════════════════════════════════════════╣
║  从 Excel 配置生成 Transformer 模型:                                 ║
║  1. xlsx 定义算子序列和量化类型                                      ║
║  2. 自动生成 Python 代码                                             ║
║  3. 运行模型并比对结果                                               ║
║                                                                      ║
║  量化策略: matmul=bfp4, 其他=bfp8                                    ║
╚══════════════════════════════════════════════════════════════════════╝
""")

    # 检查 openpyxl
    try:
        import openpyxl
    except ImportError:
        print("错误: 需要安装 openpyxl")
        print("  pip install openpyxl")
        return

    output_dir = Path(__file__).parent / "workspace"
    output_dir.mkdir(exist_ok=True)
    xlsx_path = output_dir / "transformer_config.xlsx"

    # Step 1: 创建 Transformer xlsx 配置
    print("[Step 1] 创建 Transformer xlsx 配置")
    print("-" * 50)
    ops = create_transformer_xlsx(str(xlsx_path))

    # Step 2: 显示 xlsx 内容
    print("\n[Step 2] xlsx 内容")
    print("-" * 50)
    show_xlsx_content(str(xlsx_path))

    # Step 3: 生成 Python 代码
    print("\n[Step 3] 生成 Python 代码")
    print("-" * 50)
    py_path = generate_python_code(str(xlsx_path), str(output_dir))

    # Step 4: 运行 Transformer
    print("\n[Step 4] 运行 Transformer 并比对")
    print("-" * 50)
    results = run_transformer_xlsx(str(xlsx_path), str(output_dir))

    # 总结
    print("\n" + "=" * 70)
    print("Demo 完成!")
    print("=" * 70)
    print(f"""
文件位置:
  xlsx 配置:    {xlsx_path}
  Python 代码:  {py_path}
  输出目录:     {output_dir}

量化策略:
  - linear (matmul): bfp4 (2-bit mantissa, 极端量化)
  - 其他算子:        bfp8 (4-bit mantissa, 保持精度)

模型结构:
  embedding → Q/K/V projection → attention → O projection
           → add (residual) → layernorm
           → FFN_up → gelu → FFN_down
           → add (residual) → layernorm
""")


if __name__ == "__main__":
    main()
