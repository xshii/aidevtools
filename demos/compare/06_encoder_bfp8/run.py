#!/usr/bin/env python
"""Encoder 四种比数报告 (input:bfp8, weight:bfp4)

三种前端模式 (XLSX / Torch 劫持 / DSL) 对同一 Encoder 生成 golden，
逐算子进行量化比对 (输入 bfp8, 权重 bfp4)，输出四种比数报告。

Encoder 结构:
    Q/K/V proj → Softmax → O proj → LayerNorm
    → FFN_up → GELU → FFN_down → LayerNorm

运行: python demos/compare/06_encoder_bfp8/run.py
"""
import sys
import numpy as np
from pathlib import Path

from aidevtools.datagen import DataGenerator, Model
from aidevtools.frontend.types import PrecisionConfig
from aidevtools.formats.quantize import simulate_quantize
from aidevtools.compare.fuzzy import compare_fuzzy
from aidevtools.compare.types import CompareConfig

# ---- 全局参数 ----
BATCH, SEQ, HIDDEN, FFN = 2, 16, 64, 256
QTYPE_INPUT = "bfp8"
QTYPE_WEIGHT = "bfp4"
SEED = 42

PRECISION = PrecisionConfig(
    input_dtype="fp16",
    weight_dtype="bfp4",
    compute_dtype="fp32",
    output_dtype="bfp8",
)

COMPARE_CFG = CompareConfig(
    fuzzy_min_qsnr=1.0,
    fuzzy_min_cosine=0.85,
    fuzzy_max_exceed_ratio=0.15,
    fuzzy_atol=0.5,
    fuzzy_rtol=0.5,
)

ENCODER_OPS = [
    ("Q_proj",       "linear",    (BATCH, SEQ, HIDDEN), {"out_features": HIDDEN}),
    ("K_proj",       "linear",    (BATCH, SEQ, HIDDEN), {"out_features": HIDDEN}),
    ("V_proj",       "linear",    (BATCH, SEQ, HIDDEN), {"out_features": HIDDEN}),
    ("attn_softmax", "softmax",   (BATCH, SEQ, HIDDEN), {}),
    ("O_proj",       "linear",    (BATCH, SEQ, HIDDEN), {"out_features": HIDDEN}),
    ("layernorm_1",  "layernorm", (BATCH, SEQ, HIDDEN), {}),
    ("ffn_up",       "linear",    (BATCH, SEQ, HIDDEN), {"out_features": FFN}),
    ("ffn_gelu",     "gelu",      (BATCH, SEQ, FFN),    {}),
    ("ffn_down",     "linear",    (BATCH, SEQ, FFN),    {"out_features": HIDDEN}),
    ("layernorm_2",  "layernorm", (BATCH, SEQ, HIDDEN), {}),
]


# =========================================================
# 模式 A: Model DSL
# =========================================================
def run_dsl_mode():
    """DSL 模式生成 Encoder golden (纯 Python, 无需编译器)"""
    print("\n  [A] Model DSL 模式 (纯 Python 解释执行)")

    with Model(seed=SEED, precision=PRECISION, qtype=QTYPE_INPUT) as m:
        x = m.input((BATCH, SEQ, HIDDEN))

        # Self-Attention
        q = m.linear(x, out_features=HIDDEN)
        k = m.linear(x, out_features=HIDDEN)
        v = m.linear(x, out_features=HIDDEN)
        attn = m.softmax(q)
        out = m.linear(attn, out_features=HIDDEN)
        ln1 = m.layernorm(out)

        # FFN
        up = m.linear(ln1, out_features=FFN)
        act = m.gelu(up)
        down = m.linear(act, out_features=HIDDEN)
        output = m.layernorm(down)

    golden_dsl = m.final_output.astype(np.float32) if m.final_output is not None else None
    print(f"      张量数: {len(m.tensors)}, 输出 shape: "
          f"{golden_dsl.shape if golden_dsl is not None else 'None'}")
    return golden_dsl, m


def run_dsl_four_track():
    """DSL 模式 + generate_four_track 逐算子四种比数"""
    gen = DataGenerator(seed=SEED, precision=PRECISION, qtype=QTYPE_INPUT)
    results = []

    for name, op_name, shape, kwargs in ENCODER_OPS:
        tracks = gen.generate_four_track(op_name, input_shape=shape,
                                         precision=PRECISION, **kwargs)
        results.append((name, tracks))
    return results


# =========================================================
# 模式 B: Torch 劫持
# =========================================================
def run_torch_mode():
    """Torch 劫持模式生成 Encoder golden"""
    print("\n  [B] Torch 劫持模式")

    try:
        import torch
        import torch.nn.functional as F
    except ImportError:
        print("      torch 未安装, 跳过")
        return None

    import aidevtools.golden as golden
    golden.clear()

    torch.manual_seed(SEED)

    x = torch.randn(BATCH, SEQ, HIDDEN)

    # Self-Attention
    wq = torch.randn(HIDDEN, HIDDEN) * 0.02
    wk = torch.randn(HIDDEN, HIDDEN) * 0.02
    wv = torch.randn(HIDDEN, HIDDEN) * 0.02
    q = F.linear(x, wq)
    k = F.linear(x, wk)
    v = F.linear(x, wv)
    attn = F.softmax(q, dim=-1)
    wo = torch.randn(HIDDEN, HIDDEN) * 0.02
    out = F.linear(attn, wo)
    ln1 = F.layer_norm(out, normalized_shape=(HIDDEN,))

    # FFN
    w_up = torch.randn(FFN, HIDDEN) * 0.02
    up = F.linear(ln1, w_up)
    act = F.gelu(up)
    w_down = torch.randn(HIDDEN, FFN) * 0.02
    down = F.linear(act, w_down)
    output = F.layer_norm(down, normalized_shape=(HIDDEN,))

    golden_torch = output.detach().numpy().astype(np.float32)
    records = golden.records()
    print(f"      算子记录: {len(records)} 个, 输出 shape: {golden_torch.shape}")
    return golden_torch


# =========================================================
# 模式 C: XLSX 配置
# =========================================================
def run_xlsx_mode():
    """XLSX 模式生成 Encoder golden"""
    print("\n  [C] XLSX 配置模式")

    try:
        import openpyxl
    except ImportError:
        print("      openpyxl 未安装, 跳过")
        return None

    from aidevtools.xlsx import create_template, run_xlsx
    from openpyxl import load_workbook

    workspace = Path(__file__).parent / "workspace"
    workspace.mkdir(parents=True, exist_ok=True)
    xlsx_path = workspace / "encoder_bfp4.xlsx"

    # 创建并配置模板
    create_template(str(xlsx_path), ops=["linear", "softmax", "layernorm", "gelu"])

    wb = load_workbook(xlsx_path)
    ws = wb["ops"]
    ws.delete_rows(3, ws.max_row)

    rows = [
        (0, "linear",    f"{BATCH},{SEQ},{HIDDEN}", "float32", "",  QTYPE_WEIGHT, "FALSE", "", "Q_proj"),
        (1, "linear",    f"{BATCH},{SEQ},{HIDDEN}", "float32", "",  QTYPE_WEIGHT, "FALSE", "", "K_proj"),
        (2, "linear",    f"{BATCH},{SEQ},{HIDDEN}", "float32", "",  QTYPE_WEIGHT, "FALSE", "", "V_proj"),
        (3, "softmax",   f"{BATCH},{SEQ},{HIDDEN}", "float32", "0", QTYPE_INPUT,  "FALSE", "", "softmax"),
        (4, "linear",    f"{BATCH},{SEQ},{HIDDEN}", "float32", "3", QTYPE_WEIGHT, "FALSE", "", "O_proj"),
        (5, "layernorm", f"{BATCH},{SEQ},{HIDDEN}", "float32", "4", QTYPE_INPUT,  "FALSE", "", "LN1"),
        (6, "linear",    f"{BATCH},{SEQ},{HIDDEN}", "float32", "5", QTYPE_WEIGHT, "FALSE", "", "ffn_up"),
        (7, "gelu",      f"{BATCH},{SEQ},{FFN}",    "float32", "6", QTYPE_INPUT,  "FALSE", "", "gelu"),
        (8, "linear",    f"{BATCH},{SEQ},{FFN}",    "float32", "7", QTYPE_WEIGHT, "FALSE", "", "ffn_down"),
        (9, "layernorm", f"{BATCH},{SEQ},{HIDDEN}", "float32", "8", QTYPE_INPUT,  "FALSE", "", "LN2"),
    ]
    for ri, cfg in enumerate(rows, 3):
        for ci, v in enumerate(cfg, 1):
            ws.cell(row=ri, column=ci, value=v)
    wb.save(xlsx_path)

    results = run_xlsx(str(xlsx_path), str(workspace))
    n_pass = sum(1 for r in results if r.get("status") == "PASS")
    print(f"      算子: {len(results)} 个, PASS: {n_pass}")
    return results


# =========================================================
# 比数报告
# =========================================================
def print_four_track_report(track_results):
    """打印逐算子四种比数报告"""
    print(f"\n  {'算子':<15} {'Shape':<20} "
          f"{'Pure vs HW':>12} {'Pure vs Local':>14} {'Pure vs QA':>12}")
    print(f"  {'-'*75}")

    for name, tracks in track_results:
        shape_str = str(tracks.golden_pure.shape)

        # Track 1 vs 3: pure vs hw
        hw_str = "N/A"
        if tracks.golden_hw is not None:
            r = compare_fuzzy(tracks.golden_pure, tracks.golden_hw, COMPARE_CFG)
            hw_str = f"{r.qsnr:.1f}dB"

        # Track 1 vs 2: pure vs local
        local_str = "N/A"
        if tracks.golden_local is not None:
            r = compare_fuzzy(
                tracks.golden_pure,
                tracks.golden_local.astype(np.float32) if tracks.golden_local.dtype != np.float32 else tracks.golden_local,
                COMPARE_CFG,
            )
            local_str = f"{r.qsnr:.1f}dB"

        # Track 1 vs 4: pure vs qa
        qa_str = "N/A"
        if tracks.golden_qa is not None:
            r = compare_fuzzy(tracks.golden_pure, tracks.golden_qa, COMPARE_CFG)
            qa_str = f"{r.qsnr:.1f}dB"

        print(f"  {name:<15} {shape_str:<20} {hw_str:>12} {local_str:>14} {qa_str:>12}")


def print_cross_mode_report(golden_dsl, golden_torch, xlsx_results):
    """打印跨模式一致性报告"""
    print(f"\n  跨模式最终输出量化比对 (input:{QTYPE_INPUT}, weight:{QTYPE_WEIGHT}):")
    print(f"  {'比较':<30} {'QSNR':>10} {'Cosine':>10} {'Status':>8}")
    print(f"  {'-'*60}")

    entries = []
    if golden_dsl is not None:
        dut_dsl = simulate_quantize(golden_dsl, QTYPE_INPUT)
        r = compare_fuzzy(golden_dsl, dut_dsl, COMPARE_CFG)
        entries.append((f"DSL: pure vs {QTYPE_INPUT}", r))

    if golden_torch is not None:
        dut_torch = simulate_quantize(golden_torch, QTYPE_INPUT)
        r = compare_fuzzy(golden_torch, dut_torch, COMPARE_CFG)
        entries.append((f"Torch: pure vs {QTYPE_INPUT}", r))

    if golden_dsl is not None and golden_torch is not None:
        r = compare_fuzzy(golden_dsl, golden_torch, COMPARE_CFG)
        entries.append(("DSL vs Torch (pure)", r))

    for label, r in entries:
        status = "PASS" if r.passed else "FAIL"
        print(f"  {label:<30} {r.qsnr:>10.2f} {r.cosine:>10.6f} {status:>8}")


# =========================================================
# 主函数
# =========================================================
def main():
    print("=" * 75)
    print(f"  Encoder 四种比数报告 (input:{QTYPE_INPUT}, weight:{QTYPE_WEIGHT})")
    print(f"  batch={BATCH}, seq={SEQ}, hidden={HIDDEN}, ffn={FFN}")
    print("=" * 75)

    # --- 三种模式生成 golden ---
    print("\n" + "-" * 75)
    print("  第一部分: 三种前端模式生成 Encoder golden")
    print("-" * 75)

    golden_dsl, model = run_dsl_mode()
    golden_torch = run_torch_mode()
    xlsx_results = run_xlsx_mode()

    # --- 逐算子四种比数 ---
    print("\n" + "-" * 75)
    print("  第二部分: 逐算子四种比数 (DataGenerator + PrecisionConfig)")
    print(f"  precision: input={PRECISION.input_dtype}, weight={PRECISION.weight_dtype}, "
          f"output={PRECISION.output_dtype}")
    print("-" * 75)

    track_results = run_dsl_four_track()
    print_four_track_report(track_results)

    # --- DUT 模拟 ---
    print("\n" + "-" * 75)
    print(f"  第三部分: DUT 模拟 ({QTYPE_INPUT} + noise) vs Golden")
    print("-" * 75)

    print_cross_mode_report(golden_dsl, golden_torch, xlsx_results)

    # --- 完整 DUT 模拟 ---
    print(f"\n  逐算子 DUT 模拟 ({QTYPE_INPUT} 量化 + 1% 噪声):")
    print(f"  {'算子':<15} {'QSNR':>10} {'Cosine':>10} {'MaxAbs':>12} {'Status':>8}")
    print(f"  {'-'*57}")

    rng = np.random.default_rng(SEED)
    for name, tracks in track_results:
        pure = tracks.golden_pure
        # 模拟 DUT: bfp8 量化 + 轻微噪声
        dut = simulate_quantize(pure.astype(np.float32), QTYPE_INPUT)
        noise = rng.normal(0, 0.01 * np.std(dut), size=dut.shape)
        dut = dut + noise.astype(np.float32)

        r = compare_fuzzy(pure, dut, COMPARE_CFG)
        status = "PASS" if r.passed else "FAIL"
        print(f"  {name:<15} {r.qsnr:>10.2f} {r.cosine:>10.6f} {r.max_abs:>12.2e} {status:>8}")

    print("\n" + "=" * 75)
    print("  报告完成")
    print("=" * 75)


if __name__ == "__main__":
    main()
