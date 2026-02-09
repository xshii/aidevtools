#!/usr/bin/env python
"""方式 5: Excel/XLSX 配置构建 Encoder

从 Excel 配置生成 Encoder 算子序列，输入 bfp8，权重 bfp4 精度。
非编程人员可直接在 Excel 中定义算子参数。

Encoder 结构:
    Q/K/V proj → Softmax → O proj → LayerNorm
    → FFN_up → GELU → FFN_down → LayerNorm

运行: python demos/datagen/04_xlsx_config/run.py
"""
import numpy as np
from pathlib import Path

BATCH, SEQ, HIDDEN, FFN = 2, 16, 64, 256
QTYPE_INPUT = "bfp8"
QTYPE_WEIGHT = "bfp4"


def main():
    print("=" * 70)
    print("  方式 5: Excel/XLSX 配置 (input:bfp8, weight:bfp4)")
    print("=" * 70)

    try:
        import openpyxl
    except ImportError:
        print("\n  openpyxl 未安装, 跳过本 demo")
        print("  安装: pip install openpyxl")
        print("=" * 70)
        return

    from aidevtools.xlsx import create_template, run_xlsx

    output_dir = Path(__file__).parent / "workspace"
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / "encoder_bfp4.xlsx"

    # 1. 创建模板
    print("\n  [1] 创建 XLSX 模板")
    create_template(str(xlsx_path), ops=["linear", "softmax", "layernorm", "gelu"])
    print(f"      模板: {xlsx_path}")

    # 2. 填写 Encoder 算子配置
    print("\n  [2] 填写 Encoder 算子配置")
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws_ops = wb["ops"]

    # 清空已有数据行
    ws_ops.delete_rows(3, ws_ops.max_row)

    # Encoder 算子定义: 权重用 bfp4, 输入用 bfp8
    encoder_ops = [
        # id, op_name,   shape,                 dtype,     depends, qtype,       skip,  sim_cmd, note
        (0,  "linear",   f"{BATCH},{SEQ},{HIDDEN}", "float32", "",     QTYPE_WEIGHT, "FALSE", "",  "Q_proj"),
        (1,  "linear",   f"{BATCH},{SEQ},{HIDDEN}", "float32", "",     QTYPE_WEIGHT, "FALSE", "",  "K_proj"),
        (2,  "linear",   f"{BATCH},{SEQ},{HIDDEN}", "float32", "",     QTYPE_WEIGHT, "FALSE", "",  "V_proj"),
        (3,  "softmax",  f"{BATCH},{SEQ},{HIDDEN}", "float32", "0",    QTYPE_INPUT,  "FALSE", "",  "attn_softmax"),
        (4,  "linear",   f"{BATCH},{SEQ},{HIDDEN}", "float32", "3",    QTYPE_WEIGHT, "FALSE", "",  "O_proj"),
        (5,  "layernorm",f"{BATCH},{SEQ},{HIDDEN}", "float32", "4",    QTYPE_INPUT,  "FALSE", "",  "layernorm_1"),
        (6,  "linear",   f"{BATCH},{SEQ},{HIDDEN}", "float32", "5",    QTYPE_WEIGHT, "FALSE", "",  "ffn_up"),
        (7,  "gelu",     f"{BATCH},{SEQ},{FFN}",    "float32", "6",    QTYPE_INPUT,  "FALSE", "",  "ffn_gelu"),
        (8,  "linear",   f"{BATCH},{SEQ},{FFN}",    "float32", "7",    QTYPE_WEIGHT, "FALSE", "",  "ffn_down"),
        (9,  "layernorm",f"{BATCH},{SEQ},{HIDDEN}", "float32", "8",    QTYPE_INPUT,  "FALSE", "",  "layernorm_2"),
    ]

    for row_idx, config in enumerate(encoder_ops, 3):
        for col_idx, value in enumerate(config, 1):
            ws_ops.cell(row=row_idx, column=col_idx, value=value)

    wb.save(xlsx_path)

    print(f"      {len(encoder_ops)} 个算子已配置:")
    for op in encoder_ops:
        dep = f" (depends={op[4]})" if op[4] else ""
        print(f"        [{op[0]}] {op[1]:<12} {op[8]:<15} qtype={op[5]}{dep}")

    # 3. 运行 XLSX
    print(f"\n  [3] 运行 XLSX 配置")
    results = run_xlsx(str(xlsx_path), str(output_dir))

    print(f"\n      {'ID':>4} {'算子':<12} {'Note':<15} {'Status':<8}")
    print(f"      {'-'*42}")
    for r in results:
        status_icon = "V" if r.get("status") == "PASS" else "X"
        print(f"      [{status_icon}] {r.get('id', '?'):>2} {r.get('op_name', '?'):<12} "
              f"{r.get('note', ''):<15} {r.get('status', '?'):<8}")

    print(f"\n      输出目录: {output_dir}")
    print("\n" + "=" * 70)


if __name__ == "__main__":
    main()
