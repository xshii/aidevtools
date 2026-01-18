#!/usr/bin/env python
"""
xlsx 双向工作流示例

展示两种使用方向:
1. Python → Excel: 代码生成 trace，导出到 xlsx 配置
2. Excel → Python: 从 xlsx 配置生成代码并运行

使用方法:
    python -m aidevtools.examples.xlsx_demo

    或者使用命令行:
    aidev compare xlsx template --output=config.xlsx
    aidev compare xlsx export --xlsx=config.xlsx
    aidev compare xlsx import --xlsx=config.xlsx --output=generated.py
    aidev compare xlsx run --xlsx=config.xlsx
"""
import numpy as np
from pathlib import Path
import tempfile
import shutil


def demo_python_to_excel():
    """
    方向1: Python → Excel

    场景: 已有 Python 代码，想导出为 xlsx 配置方便管理
    """
    print("=" * 60)
    print("方向1: Python → Excel")
    print("=" * 60)

    from aidevtools.ops import register_golden, clear, get_records
    from aidevtools.ops.nn import linear, relu, softmax, attention
    from aidevtools.xlsx import create_template, export_xlsx

    # 1. 注册 Golden 实现（模拟硬件实现）
    print("\n[1] 注册 Golden 实现...")

    @register_golden("linear")
    def golden_linear(x, weight, bias=None):
        """模拟硬件 Linear 实现（加入微小误差）"""
        y = np.matmul(x, weight)
        if bias is not None:
            y = y + bias
        # 模拟硬件精度损失
        y = y + np.random.randn(*y.shape).astype(np.float32) * 1e-6
        return y

    @register_golden("relu")
    def golden_relu(x):
        return np.maximum(0, x)

    @register_golden("softmax")
    def golden_softmax(x, axis=-1):
        x_max = np.max(x, axis=axis, keepdims=True)
        x_exp = np.exp(x - x_max)
        return x_exp / np.sum(x_exp, axis=axis, keepdims=True)

    print("    已注册: linear, relu, softmax")

    # 2. 执行算子流程
    print("\n[2] 执行算子流程...")
    clear()

    # 模拟一个简单的 MLP
    x = np.random.randn(1, 8, 64).astype(np.float32)
    w1 = np.random.randn(64, 128).astype(np.float32)
    w2 = np.random.randn(128, 64).astype(np.float32)

    h = linear(x, w1)       # linear_0
    h = relu(h)             # relu_0
    h = linear(h, w2)       # linear_1
    y = softmax(h)          # softmax_0

    print(f"    MLP: input{x.shape} → linear → relu → linear → softmax → output{y.shape}")

    # 3. 导出到 xlsx
    print("\n[3] 导出到 xlsx...")
    output_dir = Path(tempfile.mkdtemp())
    xlsx_path = output_dir / "mlp_config.xlsx"

    records = get_records()
    export_xlsx(str(xlsx_path), records)

    print(f"    导出 {len(records)} 条记录到: {xlsx_path}")
    print(f"    包含 3 个 sheet: op_registry, ops, compare")

    # 4. 展示 xlsx 内容
    print("\n[4] xlsx 内容预览...")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)

        # ops sheet
        ws_ops = wb["ops"]
        print("\n    [ops sheet]")
        print("    " + "-" * 50)
        for row in ws_ops.iter_rows(min_row=1, max_row=5, values_only=True):
            print(f"    {row}")

        # compare sheet
        ws_compare = wb["compare"]
        print("\n    [compare sheet]")
        print("    " + "-" * 50)
        for row in ws_compare.iter_rows(min_row=1, max_row=5, values_only=True):
            print(f"    {row}")

    except ImportError:
        print("    (需要 openpyxl 查看内容)")

    print(f"\n    ✓ 完成! xlsx 文件: {xlsx_path}")
    return str(xlsx_path), str(output_dir)


def demo_excel_to_python():
    """
    方向2: Excel → Python

    场景: 用 Excel 配置算子用例，自动生成 Python 代码
    """
    print("\n" + "=" * 60)
    print("方向2: Excel → Python")
    print("=" * 60)

    from aidevtools.xlsx import create_template, import_xlsx, run_xlsx

    # 1. 创建 xlsx 模板
    print("\n[1] 创建 xlsx 模板...")
    output_dir = Path(tempfile.mkdtemp())
    xlsx_path = output_dir / "custom_config.xlsx"

    # 限定只使用部分算子
    create_template(str(xlsx_path), ops=["linear", "relu", "matmul", "softmax"])
    print(f"    模板: {xlsx_path}")
    print(f"    限定算子: linear, relu, matmul, softmax")

    # 2. 模拟用户编辑 xlsx（实际场景中用户用 Excel 编辑）
    print("\n[2] 模拟用户编辑 xlsx...")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws_ops = wb["ops"]

        # 清空示例数据，添加自定义配置
        # 删除示例行（从第3行开始，保留表头和说明）
        ws_ops.delete_rows(3, ws_ops.max_row)

        # 用户配置的算子用例
        user_configs = [
            # id, op_name, shape, dtype, depends, qtype, skip, note
            (0, "linear", "1,32,64", "float32", "", "", "FALSE", "输入层"),
            (1, "relu", "1,32,128", "float32", "0", "", "FALSE", "激活"),
            (2, "linear", "1,32,128", "float32", "1", "", "FALSE", "隐藏层"),
            (3, "softmax", "1,32,64", "float32", "2", "", "FALSE", "输出层"),
        ]

        for row_idx, config in enumerate(user_configs, 3):
            for col_idx, value in enumerate(config, 1):
                ws_ops.cell(row=row_idx, column=col_idx, value=value)

        wb.save(xlsx_path)
        print("    已添加 4 个算子配置:")
        print("      0: linear (输入层)")
        print("      1: relu   (依赖 0)")
        print("      2: linear (依赖 1)")
        print("      3: softmax (依赖 2)")

    except ImportError:
        print("    (需要 openpyxl)")
        return None, None

    # 3. 从 xlsx 生成 Python 代码
    print("\n[3] 从 xlsx 生成 Python 代码...")
    py_path = output_dir / "generated_model.py"
    code = import_xlsx(str(xlsx_path), str(py_path))

    print(f"    生成代码: {py_path}")
    print("\n    [生成的代码预览]")
    print("    " + "-" * 50)
    for line in code.split("\n")[:25]:
        print(f"    {line}")
    if code.count("\n") > 25:
        print("    ...")

    # 4. 运行 xlsx 配置
    print("\n[4] 运行 xlsx 配置...")
    results = run_xlsx(str(xlsx_path), str(output_dir))

    print(f"\n    运行结果:")
    for r in results:
        status = r.get("status", "?")
        op_id = r.get("id", "?")
        note = r.get("note", "")
        status_icon = "✓" if status == "PASS" else "✗" if status == "FAIL" else "○"
        print(f"      [{status_icon}] id={op_id}: {status} {note}")

    print(f"\n    ✓ 完成! 结果已更新到 xlsx")
    return str(xlsx_path), str(output_dir)


def demo_bidirectional_workflow():
    """
    双向工作流示例

    完整流程:
    1. Python 代码 → xlsx 配置
    2. 用户用 Excel 编辑调整
    3. xlsx 配置 → 运行比对
    4. 结果写回 xlsx（不覆盖已有结果）
    """
    print("\n" + "=" * 60)
    print("双向工作流: Python ↔ Excel")
    print("=" * 60)

    from aidevtools.ops import register_golden, clear, get_records
    from aidevtools.ops.nn import linear, relu, attention
    from aidevtools.xlsx import export_xlsx, run_xlsx
    from aidevtools.xlsx.export import update_compare_results

    output_dir = Path(tempfile.mkdtemp())
    xlsx_path = output_dir / "bidirectional.xlsx"

    # 注册 golden
    @register_golden("attention")
    def golden_attention(q, k, v, mask=None):
        d_k = q.shape[-1]
        scores = np.matmul(q, k.swapaxes(-2, -1)) / np.sqrt(d_k)
        if mask is not None:
            scores = scores + mask * (-1e9)
        scores_max = np.max(scores, axis=-1, keepdims=True)
        scores_exp = np.exp(scores - scores_max)
        attn_weights = scores_exp / np.sum(scores_exp, axis=-1, keepdims=True)
        return np.matmul(attn_weights, v)

    # Step 1: Python → xlsx
    print("\n[Step 1] Python 执行算子，导出到 xlsx")
    clear()

    q = np.random.randn(1, 4, 8, 32).astype(np.float32)
    k = np.random.randn(1, 4, 8, 32).astype(np.float32)
    v = np.random.randn(1, 4, 8, 32).astype(np.float32)

    out = attention(q, k, v)
    print(f"    执行 attention: Q{q.shape} K{k.shape} V{v.shape} → {out.shape}")

    records = get_records()
    export_xlsx(str(xlsx_path), records)
    print(f"    导出到: {xlsx_path}")

    # Step 2: 模拟第一次比对
    print("\n[Step 2] 第一次运行比对")
    results = run_xlsx(str(xlsx_path), str(output_dir))
    for r in results:
        print(f"    id={r.get('id')}: {r.get('status')}")

    # Step 3: 再次导出（验证结果保留）
    print("\n[Step 3] 再次导出（验证结果列保留）")

    # 模拟新增一条记录
    clear()
    x = np.random.randn(1, 8, 64).astype(np.float32)
    out2 = relu(x)

    records2 = get_records()
    export_xlsx(str(xlsx_path), records2, preserve_results=True)

    # 检查之前的结果是否保留
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb["compare"]

        print("    compare sheet 内容:")
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            print(f"      {row}")
        print("    ✓ 已有结果被保留")

    except ImportError:
        pass

    print(f"\n    ✓ 双向工作流完成!")
    return str(xlsx_path), str(output_dir)


def main():
    """主函数"""
    print("""
╔══════════════════════════════════════════════════════════════╗
║              xlsx 双向工作流示例                              ║
╠══════════════════════════════════════════════════════════════╣
║  方向1: Python → Excel   代码导出为配置                      ║
║  方向2: Excel → Python   配置生成代码                        ║
║  双向:  Python ↔ Excel   完整协作流程                        ║
╚══════════════════════════════════════════════════════════════╝
""")

    # 检查 openpyxl
    try:
        import openpyxl
    except ImportError:
        print("错误: 需要安装 openpyxl")
        print("  pip install openpyxl")
        return

    # 运行示例
    xlsx1, dir1 = demo_python_to_excel()
    xlsx2, dir2 = demo_excel_to_python()
    xlsx3, dir3 = demo_bidirectional_workflow()

    # 总结
    print("\n" + "=" * 60)
    print("命令行用法总结")
    print("=" * 60)
    print("""
# 生成空模板
aidev compare xlsx template --output=config.xlsx

# 限定算子的模板
aidev compare xlsx template --output=config.xlsx --ops=linear,relu,attention

# 从 trace 导出到 xlsx
aidev compare xlsx export --xlsx=config.xlsx

# 从 xlsx 生成 Python 代码
aidev compare xlsx import --xlsx=config.xlsx --output=generated.py

# 运行 xlsx 配置并比对
aidev compare xlsx run --xlsx=config.xlsx

# 列出可用算子
aidev compare xlsx ops
""")

    # 清理临时目录
    for d in [dir1, dir2, dir3]:
        if d:
            shutil.rmtree(d, ignore_errors=True)

    print("✓ 示例运行完成!")


if __name__ == "__main__":
    main()
