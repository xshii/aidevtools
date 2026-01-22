"""Export 模块 - Excel 导出与 Gantt 图

功能:
- 导出时延分析结果到 xlsx
- 生成流水图 (Gantt Chart) 页签
- 支持多种格式化选项
"""

from pathlib import Path
from typing import Optional, List, Dict, Any
from dataclasses import dataclass

from .latency import LatencyResult, LatencyBreakdown, GanttData, GanttItem
from .analyzer import AnalysisSummary


# Gantt 图颜色映射
GANTT_COLORS = {
    "cube": "4472C4",       # 蓝色
    "vector": "70AD47",     # 绿色
    "dma": "FFC000",        # 黄色
    "execution": "4472C4",  # 蓝色
    "prefetch": "FFC000",   # 黄色
    "parallel": "7030A0",   # 紫色
}


def export_xlsx(result: LatencyResult,
                output_path: str,
                include_gantt: bool = True,
                include_passes: bool = True,
                include_summary: bool = True):
    """导出分析结果到 xlsx

    Args:
        result: 分析结果
        output_path: 输出文件路径
        include_gantt: 是否包含 Gantt 图页签
        include_passes: 是否包含 Pass 详情页签
        include_summary: 是否包含摘要页签
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    except ImportError:
        raise ImportError("openpyxl is required for xlsx export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. 主分析结果页签
    ws_main = wb.active
    ws_main.title = "Latency Analysis"
    _write_main_sheet(ws_main, result, header_font, header_fill, border)

    # 2. 摘要页签
    if include_summary and result.summary:
        ws_summary = wb.create_sheet("Summary")
        _write_summary_sheet(ws_summary, result, header_font, header_fill, border)

    # 3. Pass 详情页签
    if include_passes and result.pass_results:
        ws_passes = wb.create_sheet("Pass Details")
        _write_passes_sheet(ws_passes, result, header_font, header_fill, border)

    # 4. Gantt 图页签
    if include_gantt and result.gantt_data:
        ws_gantt = wb.create_sheet("Gantt Chart")
        _write_gantt_sheet(ws_gantt, result.gantt_data, header_font, header_fill, border)

    # 保存文件
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Exported to: {output_path}")


def _write_main_sheet(ws, result: LatencyResult, header_font, header_fill, border):
    """写入主分析结果页签"""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    # 表头
    headers = [
        "Op Name", "Op Type", "Compute Unit", "Dtype",
        "FLOPs (M)", "Input (KB)", "Weight (KB)", "Output (KB)",
        "Compute (us)", "Memory (us)", "Roofline (us)",
        "Prefetch Saved", "Parallel Saved", "Overhead",
        "Total (us)", "Bottleneck", "Min BW (GB/s)"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # 数据行
    for row_idx, bd in enumerate(result.breakdowns, 2):
        p = bd.profile
        values = [
            p.name,
            p.op_type,
            p.compute_unit,
            p.dtype,
            p.flops / 1e6,
            p.input_bytes / 1024,
            p.weight_bytes / 1024,
            p.output_bytes / 1024,
            bd.compute_time_us,
            bd.memory_time_us,
            bd.roofline_time_us,
            bd.prefetch_saved_us,
            bd.parallel_saved_us,
            bd.overhead_us,
            bd.total_time_us,
            bd.bottleneck,
            bd.min_bandwidth_gbps,
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if isinstance(value, float):
                cell.number_format = '0.00'

    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def _write_summary_sheet(ws, result: LatencyResult, header_font, header_fill, border):
    """写入摘要页签"""
    from openpyxl.styles import Font

    s = result.summary
    chip = result.chip_spec

    # 标题
    ws.cell(row=1, column=1, value=f"Paper Analysis Summary - {chip.name}").font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    # 基本信息
    data = [
        ("", ""),
        ("Chip Info", ""),
        ("Chip Name", chip.name),
        ("Cube FP16 TFLOPS", chip.cube.fp16_tflops),
        ("Vector FP16 GFLOPS", chip.vector.fp16_gflops),
        ("HBM Bandwidth (GB/s)", chip.memory.hbm.bandwidth_gbps),
        ("HBM Capacity (GB)", chip.memory.hbm.capacity_bytes / 1024**3),
        ("", ""),
        ("Analysis Summary", ""),
        ("Total Operators", len(result.breakdowns)),
        ("Total Latency (us)", s.total_latency_us),
        ("Total Latency (ms)", s.total_latency_us / 1000),
        ("", ""),
        ("Time Breakdown", ""),
        ("Compute Time (us)", s.total_compute_time_us),
        ("Memory Time (us)", s.total_memory_time_us),
        ("Overhead (us)", s.total_overhead_us),
        ("", ""),
        ("Bottleneck Stats", ""),
        ("Compute Bound Ops", s.compute_bound_ops),
        ("Memory Bound Ops", s.memory_bound_ops),
        ("", ""),
        ("Optimizations", ""),
        ("Prefetch Saved (us)", s.total_prefetch_saved_us),
        ("Parallel Saved (us)", s.total_parallel_saved_us),
        ("", ""),
        ("Throughput", ""),
        ("Achieved TFLOPS", s.achieved_tflops),
        ("Achieved Bandwidth (GB/s)", s.achieved_bandwidth_gbps),
        ("", ""),
        ("Unit Utilization", ""),
        ("Cube Time (us)", s.cube_time_us),
        ("Vector Time (us)", s.vector_time_us),
        ("Cube Ratio (%)", s.cube_time_us / s.total_latency_us * 100 if s.total_latency_us > 0 else 0),
        ("Vector Ratio (%)", s.vector_time_us / s.total_latency_us * 100 if s.total_latency_us > 0 else 0),
    ]

    for row_idx, (label, value) in enumerate(data, 3):
        ws.cell(row=row_idx, column=1, value=label)
        if value != "":
            cell = ws.cell(row=row_idx, column=2, value=value)
            if isinstance(value, float):
                cell.number_format = '0.00'

        # 小节标题加粗
        if label and value == "":
            ws.cell(row=row_idx, column=1).font = Font(bold=True)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15


def _write_passes_sheet(ws, result: LatencyResult, header_font, header_fill, border):
    """写入 Pass 详情页签"""
    from openpyxl.utils import get_column_letter

    # 表头
    headers = ["Op Name", "Pass Name", "Enabled", "Before (us)", "After (us)", "Saved (us)", "Improvement %"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    row_idx = 2
    for i, (bd, passes) in enumerate(zip(result.breakdowns, result.pass_results)):
        for pr in passes:
            values = [
                bd.profile.name,
                pr.pass_name,
                "Yes" if pr.enabled else "No",
                pr.latency_before_us,
                pr.latency_after_us,
                pr.latency_saved_us,
                pr.improvement_ratio * 100 if pr.latency_before_us > 0 else 0,
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if isinstance(value, float):
                    cell.number_format = '0.00'

            row_idx += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _write_gantt_sheet(ws, gantt_data: GanttData, header_font, header_fill, border):
    """写入 Gantt 图页签

    使用条件格式和单元格着色模拟 Gantt 图
    """
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    # 标题
    ws.cell(row=1, column=1, value=f"Pipeline Gantt Chart - {gantt_data.chip_name}").font = Font(bold=True, size=14)
    ws.merge_cells('A1:Z1')

    # 时间刻度 (每列代表一定时间)
    time_scale_us = gantt_data.total_time_us / 50 if gantt_data.total_time_us > 0 else 1  # 50 列
    ws.cell(row=2, column=1, value=f"Time Scale: {time_scale_us:.2f} us/column")

    # 图例
    row = 4
    ws.cell(row=row, column=1, value="Legend:")
    row += 1

    for unit, color in [("Cube", "4472C4"), ("Vector", "70AD47"), ("DMA/Prefetch", "FFC000")]:
        ws.cell(row=row, column=1, value=unit)
        cell = ws.cell(row=row, column=2, value="")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    # Gantt 表头
    row += 2
    ws.cell(row=row, column=1, value="Op Name").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Unit").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Start (us)").font = Font(bold=True)
    ws.cell(row=row, column=4, value="End (us)").font = Font(bold=True)
    ws.cell(row=row, column=5, value="Duration (us)").font = Font(bold=True)

    # 时间轴标题 (从第6列开始)
    for i in range(50):
        col = 6 + i
        time_val = i * time_scale_us
        if i % 10 == 0:
            ws.cell(row=row, column=col, value=f"{time_val:.0f}")
        ws.column_dimensions[get_column_letter(col)].width = 2

    header_row = row
    row += 1

    # Gantt 条目
    for item in gantt_data.items:
        ws.cell(row=row, column=1, value=item.op_name)
        ws.cell(row=row, column=2, value=item.unit)
        ws.cell(row=row, column=3, value=item.start_us).number_format = '0.00'
        ws.cell(row=row, column=4, value=item.end_us).number_format = '0.00'
        ws.cell(row=row, column=5, value=item.end_us - item.start_us).number_format = '0.00'

        # 绘制 Gantt 条
        start_col = int(item.start_us / time_scale_us) + 6 if time_scale_us > 0 else 6
        end_col = int(item.end_us / time_scale_us) + 6 if time_scale_us > 0 else 6

        # 选择颜色
        color = GANTT_COLORS.get(item.unit, GANTT_COLORS.get(item.category, "808080"))

        for col in range(start_col, min(end_col + 1, 56)):  # 最多50列
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        row += 1

    # 总时间
    row += 1
    ws.cell(row=row, column=1, value="Total Time (us):").font = Font(bold=True)
    ws.cell(row=row, column=2, value=gantt_data.total_time_us).number_format = '0.00'

    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12


def export_csv(result: LatencyResult, output_path: str):
    """导出为 CSV 格式"""
    import csv
    from pathlib import Path

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # 表头
        headers = [
            "Op Name", "Op Type", "Compute Unit", "Dtype",
            "FLOPs", "Input Bytes", "Weight Bytes", "Output Bytes",
            "Compute Time (us)", "Memory Time (us)", "Roofline Time (us)",
            "Prefetch Saved (us)", "Parallel Saved (us)", "Overhead (us)",
            "Total Time (us)", "Bottleneck", "Min Bandwidth (GB/s)"
        ]
        writer.writerow(headers)

        # 数据行
        for bd in result.breakdowns:
            p = bd.profile
            row = [
                p.name, p.op_type, p.compute_unit, p.dtype,
                p.flops, p.input_bytes, p.weight_bytes, p.output_bytes,
                bd.compute_time_us, bd.memory_time_us, bd.roofline_time_us,
                bd.prefetch_saved_us, bd.parallel_saved_us, bd.overhead_us,
                bd.total_time_us, bd.bottleneck, bd.min_bandwidth_gbps
            ]
            writer.writerow(row)

    print(f"Exported to: {output_path}")


def export_json(result: LatencyResult, output_path: str):
    """导出为 JSON 格式"""
    import json
    from pathlib import Path
    from dataclasses import asdict

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 构建可序列化的数据结构
    data = {
        "chip": {
            "name": result.chip_spec.name,
            "cube_fp16_tflops": result.chip_spec.cube.fp16_tflops,
            "vector_fp16_gflops": result.chip_spec.vector.fp16_gflops,
            "hbm_bandwidth_gbps": result.chip_spec.memory.hbm.bandwidth_gbps,
        },
        "summary": {
            "total_latency_us": result.summary.total_latency_us,
            "total_flops": result.summary.total_flops,
            "total_bytes": result.summary.total_bytes,
            "compute_bound_ops": result.summary.compute_bound_ops,
            "memory_bound_ops": result.summary.memory_bound_ops,
            "achieved_tflops": result.summary.achieved_tflops,
            "achieved_bandwidth_gbps": result.summary.achieved_bandwidth_gbps,
        },
        "breakdowns": []
    }

    for bd in result.breakdowns:
        p = bd.profile
        breakdown_data = {
            "op_name": p.name,
            "op_type": p.op_type,
            "compute_unit": p.compute_unit,
            "dtype": p.dtype,
            "flops": p.flops,
            "input_bytes": p.input_bytes,
            "weight_bytes": p.weight_bytes,
            "output_bytes": p.output_bytes,
            "compute_time_us": bd.compute_time_us,
            "memory_time_us": bd.memory_time_us,
            "roofline_time_us": bd.roofline_time_us,
            "total_time_us": bd.total_time_us,
            "bottleneck": bd.bottleneck,
            "min_bandwidth_gbps": bd.min_bandwidth_gbps,
        }
        data["breakdowns"].append(breakdown_data)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Exported to: {output_path}")
