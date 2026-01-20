"""xlsx 导出

从 trace 记录导出到 xlsx，保留已有的结果列。
"""
from pathlib import Path
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from aidevtools.core.log import logger


def _check_openpyxl():
    if not HAS_OPENPYXL:
        raise ImportError("xlsx 功能需要 openpyxl，请安装: pip install openpyxl")


def export_xlsx(
    output_path: str,
    records: List[Dict[str, Any]],
    preserve_results: bool = True,
) -> str:
    """
    导出 trace 记录到 xlsx

    Args:
        output_path: 输出文件路径
        records: trace 记录列表
        preserve_results: 是否保留已有的结果列 (compare sheet)

    Returns:
        输出文件路径

    重要:
        - 如果文件已存在且 preserve_results=True，会保留 compare sheet 中的结果数据
        - ops sheet 会被更新为最新的记录
    """
    _check_openpyxl()

    output_path = Path(output_path)
    existing_compare_data = {}

    # 读取已有的 compare 结果
    if preserve_results and output_path.exists():
        try:
            existing_wb = load_workbook(output_path)
            if "compare" in existing_wb.sheetnames:
                ws = existing_wb["compare"]
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    if row_dict.get("id") is not None:
                        existing_compare_data[row_dict["id"]] = row_dict
                logger.debug(f"保留 {len(existing_compare_data)} 条已有比对结果")
            existing_wb.close()
        except Exception as e:
            logger.warning(f"读取已有结果失败: {e}")

    # 创建或加载工作簿
    if output_path.exists():
        wb = load_workbook(output_path)
    else:
        from aidevtools.xlsx.template import create_template
        create_template(str(output_path), include_examples=False)
        wb = load_workbook(output_path)

    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==================== 更新 ops sheet ====================
    if "ops" in wb.sheetnames:
        ws_ops = wb["ops"]
        # 清空数据行 (保留表头)
        ws_ops.delete_rows(2, ws_ops.max_row)
    else:
        ws_ops = wb.create_sheet("ops")
        ops_headers = ["id", "op_name", "shape", "dtype", "depends", "qtype", "skip", "note"]
        for col, header in enumerate(ops_headers, 1):
            cell = ws_ops.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

    # 写入记录
    import numpy as np
    for idx, record in enumerate(records):
        row = idx + 2
        name = record.get("name", f"op_{idx}")
        op = record.get("op", name.rsplit("_", 1)[0])

        # 获取 shape 和 dtype
        output = record.get("golden")
        if output is None:
            output = record.get("output")
        if output is not None and hasattr(output, 'shape'):
            arr = np.asarray(output)
            shape_str = ",".join(str(d) for d in arr.shape)
            dtype_str = str(arr.dtype)
        else:
            shape_str = ""
            dtype_str = ""

        ws_ops.cell(row=row, column=1, value=idx).border = thin_border
        ws_ops.cell(row=row, column=2, value=op).border = thin_border
        ws_ops.cell(row=row, column=3, value=shape_str).border = thin_border
        ws_ops.cell(row=row, column=4, value=dtype_str).border = thin_border
        ws_ops.cell(row=row, column=5, value="").border = thin_border  # depends 由用户填写
        ws_ops.cell(row=row, column=6, value="").border = thin_border  # qtype
        ws_ops.cell(row=row, column=7, value="FALSE").border = thin_border  # skip
        ws_ops.cell(row=row, column=8, value=name).border = thin_border  # note 中放 full name

    # ==================== 更新 compare sheet (保留结果) ====================
    if "compare" in wb.sheetnames:
        ws_compare = wb["compare"]
        ws_compare.delete_rows(2, ws_compare.max_row)
    else:
        ws_compare = wb.create_sheet("compare")
        compare_headers = ["id", "op_name", "status", "max_abs", "qsnr", "cosine", "golden_bin", "result_bin", "note"]
        for col, header in enumerate(compare_headers, 1):
            cell = ws_compare.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

    # 写入 compare 数据 (合并已有结果)
    for idx, record in enumerate(records):
        row = idx + 2
        name = record.get("name", f"op_{idx}")

        # 检查是否有已保存的结果
        existing = existing_compare_data.get(idx, {})

        ws_compare.cell(row=row, column=1, value=idx).border = thin_border
        ws_compare.cell(row=row, column=2, value=name).border = thin_border

        # 保留已有结果，否则留空
        ws_compare.cell(row=row, column=3, value=existing.get("status", "")).border = thin_border
        ws_compare.cell(row=row, column=4, value=existing.get("max_abs", "")).border = thin_border
        ws_compare.cell(row=row, column=5, value=existing.get("qsnr", "")).border = thin_border
        ws_compare.cell(row=row, column=6, value=existing.get("cosine", "")).border = thin_border
        ws_compare.cell(row=row, column=7, value=existing.get("golden_bin", "")).border = thin_border
        ws_compare.cell(row=row, column=8, value=existing.get("result_bin", "")).border = thin_border
        ws_compare.cell(row=row, column=9, value=existing.get("note", "")).border = thin_border

    # 保存
    wb.save(output_path)
    logger.info(f"导出 xlsx: {output_path} ({len(records)} 条记录)")
    return str(output_path)


def update_compare_results(
    xlsx_path: str,
    results: List[Dict[str, Any]],
) -> str:
    """
    更新 xlsx 中的比对结果

    Args:
        xlsx_path: xlsx 文件路径
        results: 比对结果列表，每项包含 id, status, max_abs, qsnr, cosine 等

    Returns:
        xlsx 文件路径
    """
    _check_openpyxl()

    wb = load_workbook(xlsx_path)
    if "compare" not in wb.sheetnames:
        raise ValueError(f"xlsx 文件缺少 compare sheet: {xlsx_path}")

    ws = wb["compare"]

    # 读取表头
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers)}

    # 构建 id -> row 映射
    id_to_row = {}
    for row in range(2, ws.max_row + 1):
        cell_id = ws.cell(row=row, column=col_map.get("id", 1)).value
        if cell_id is not None:
            id_to_row[cell_id] = row

    # 结果样式
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 更新结果
    for res in results:
        res_id = res.get("id")
        if res_id not in id_to_row:
            continue

        row = id_to_row[res_id]
        status = res.get("status", "")

        if "status" in col_map:
            cell = ws.cell(row=row, column=col_map["status"], value=status)
            if status == "PASS":
                cell.fill = pass_fill
            elif status == "FAIL":
                cell.fill = fail_fill

        if "max_abs" in col_map and "max_abs" in res:
            ws.cell(row=row, column=col_map["max_abs"], value=res["max_abs"])
        if "qsnr" in col_map and "qsnr" in res:
            ws.cell(row=row, column=col_map["qsnr"], value=res["qsnr"])
        if "cosine" in col_map and "cosine" in res:
            ws.cell(row=row, column=col_map["cosine"], value=res["cosine"])
        if "golden_bin" in col_map and "golden_bin" in res:
            ws.cell(row=row, column=col_map["golden_bin"], value=res["golden_bin"])
        if "result_bin" in col_map and "result_bin" in res:
            ws.cell(row=row, column=col_map["result_bin"], value=res["result_bin"])
        if "note" in col_map and "note" in res:
            ws.cell(row=row, column=col_map["note"], value=res["note"])

        # IsClose 比对结果
        if "isclose_pass" in col_map and "isclose_pass" in res:
            cell = ws.cell(row=row, column=col_map["isclose_pass"], value=res["isclose_pass"])
            if res["isclose_pass"] == "PASS":
                cell.fill = pass_fill
            elif res["isclose_pass"] == "FAIL":
                cell.fill = fail_fill
        if "exceed_count" in col_map and "exceed_count" in res:
            ws.cell(row=row, column=col_map["exceed_count"], value=res["exceed_count"])
        if "exceed_ratio" in col_map and "exceed_ratio" in res:
            ws.cell(row=row, column=col_map["exceed_ratio"], value=res["exceed_ratio"])

    wb.save(xlsx_path)
    logger.info(f"更新比对结果: {xlsx_path}")
    return xlsx_path
