"""xlsx 导入

从 xlsx 解析配置并生成 Python 代码。
"""
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from dataclasses import dataclass

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from aidevtools.core.log import logger


def _check_openpyxl():
    if not HAS_OPENPYXL:
        raise ImportError("xlsx 功能需要 openpyxl，请安装: pip install openpyxl")


@dataclass
class OpConfig:
    """算子配置"""
    id: int
    op_name: str
    shape: Tuple[int, ...]
    dtype: str
    depends: str  # 原始依赖字符串
    qtype: str
    skip: bool
    note: str
    sim_cmd: str = ""  # 仿真命令，支持占位符: {golden_bin}, {result_bin}, {input_bin}, {weight_bin}, {id}, {op_name}
    # binary 路径（留空=自动生成，填写=使用指定路径）
    golden_bin: str = ""
    result_bin: str = ""
    input_bin: str = ""
    weight_bin: str = ""

    def parse_depends(self) -> Dict[str, List[int]]:
        """
        解析依赖关系

        Returns:
            依赖映射，格式: {"input_name": [row_ids]}

        示例:
            "" -> {}  (无依赖，随机输入)
            "0" -> {"x": [0]}  (单依赖)
            "1,2" -> {"a": [1], "b": [2]}  (双输入依赖)
            "q:0,k:1,v:2" -> {"q": [0], "k": [1], "v": [2]}  (命名依赖)
        """
        if not self.depends or self.depends.strip() == "":
            return {}

        depends_str = self.depends.strip()
        result = {}

        # 检查是否是命名依赖 (包含 ":")
        if ":" in depends_str:
            # 命名依赖: "q:0,k:1,v:2"
            parts = depends_str.split(",")
            for part in parts:
                if ":" not in part:
                    continue
                name, idx_str = part.split(":", 1)
                name = name.strip()
                try:
                    idx = int(idx_str.strip())
                    result[name] = [idx]
                except ValueError:
                    logger.warn(f"无效的依赖索引: {part}")
        else:
            # 简单依赖: "0" 或 "1,2"
            parts = [p.strip() for p in depends_str.split(",") if p.strip()]
            if len(parts) == 1:
                # 单依赖
                try:
                    result["x"] = [int(parts[0])]
                except ValueError:
                    logger.warn(f"无效的依赖索引: {parts[0]}")
            else:
                # 双输入依赖
                input_names = ["a", "b", "c", "d", "e", "f"]
                for i, part in enumerate(parts):
                    try:
                        name = input_names[i] if i < len(input_names) else f"in{i}"
                        result[name] = [int(part)]
                    except ValueError:
                        logger.warn(f"无效的依赖索引: {part}")

        return result


def parse_xlsx(xlsx_path: str) -> Tuple[List[str], List[OpConfig]]:
    """
    解析 xlsx 文件

    Args:
        xlsx_path: xlsx 文件路径

    Returns:
        (enabled_ops, op_configs)
        - enabled_ops: 启用的算子列表
        - op_configs: 算子配置列表
    """
    _check_openpyxl()

    wb = load_workbook(xlsx_path, data_only=True)

    # 解析 op_registry sheet
    enabled_ops = []
    if "op_registry" in wb.sheetnames:
        ws = wb["op_registry"]
        headers = [cell.value for cell in ws[1]]
        col_map = {h: i for i, h in enumerate(headers)}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            row_dict = dict(zip(headers, row))
            op_name = row_dict.get("op_name", "")
            enabled = str(row_dict.get("enabled", "TRUE")).upper()
            if op_name and enabled == "TRUE":
                enabled_ops.append(op_name)

    # 解析 ops sheet
    op_configs = []
    if "ops" in wb.sheetnames:
        ws = wb["ops"]
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            # 跳过注释行
            first_cell = str(row[0]).strip()
            if first_cell.startswith("#"):
                continue

            row_dict = dict(zip(headers, row))

            # 解析 shape
            shape_str = str(row_dict.get("shape", "") or "")
            try:
                shape = tuple(int(x.strip()) for x in shape_str.split(",") if x.strip())
            except ValueError:
                shape = ()

            # 解析 skip
            skip_str = str(row_dict.get("skip", "FALSE") or "FALSE").upper()
            skip = skip_str in ("TRUE", "1", "YES")

            config = OpConfig(
                id=int(row_dict.get("id", 0) or 0),
                op_name=str(row_dict.get("op_name", "") or ""),
                shape=shape,
                dtype=str(row_dict.get("dtype", "float32") or "float32"),
                depends=str(row_dict.get("depends", "") or ""),
                qtype=str(row_dict.get("qtype", "") or ""),
                skip=skip,
                note=str(row_dict.get("note", "") or ""),
                sim_cmd=str(row_dict.get("sim_cmd", "") or ""),
                golden_bin=str(row_dict.get("golden_bin", "") or ""),
                result_bin=str(row_dict.get("result_bin", "") or ""),
                input_bin=str(row_dict.get("input_bin", "") or ""),
                weight_bin=str(row_dict.get("weight_bin", "") or ""),
            )
            op_configs.append(config)

    wb.close()
    return enabled_ops, op_configs


def import_xlsx(xlsx_path: str, output_py: Optional[str] = None) -> str:
    """
    从 xlsx 生成 Python 代码

    Args:
        xlsx_path: xlsx 文件路径
        output_py: 输出 Python 文件路径，None 表示只返回代码不写文件

    Returns:
        生成的 Python 代码
    """
    enabled_ops, op_configs = parse_xlsx(xlsx_path)

    # 生成代码
    lines = [
        '"""自动生成的算子测试代码',
        f'',
        f'从 xlsx 配置生成: {Path(xlsx_path).name}',
        '"""',
        'import numpy as np',
        'from aidevtools.ops import nn',
        'from aidevtools.ops.base import clear, dump, gen_csv',
        '',
        '',
        'def run():',
        '    """执行算子测试"""',
        '    clear()  # 清空之前的记录',
        '    outputs = {}  # 保存各步骤的输出',
        '',
    ]

    for config in op_configs:
        if config.skip:
            lines.append(f'    # [SKIP] {config.op_name} (id={config.id})')
            continue

        indent = "    "
        op_name = config.op_name
        shape = config.shape
        dtype = config.dtype

        # 解析依赖
        depends = config.parse_depends()

        # 生成注释
        if config.note:
            lines.append(f'{indent}# {config.note}')

        # 生成输入
        if not depends:
            # 无依赖，随机输入
            shape_str = ", ".join(str(d) for d in shape) if shape else "1, 64"
            lines.append(f'{indent}x_{config.id} = np.random.randn({shape_str}).astype(np.{dtype})')
            input_var = f"x_{config.id}"
        else:
            # 有依赖
            input_vars = []
            for name, deps in depends.items():
                for dep_id in deps:
                    lines.append(f'{indent}{name}_{config.id} = outputs[{dep_id}]')
                    input_vars.append(f"{name}_{config.id}")
            input_var = ", ".join(input_vars) if len(input_vars) > 1 else input_vars[0] if input_vars else f"x_{config.id}"

        # 生成算子调用
        if op_name == "linear":
            if not depends:
                shape_str = ", ".join(str(d) for d in shape) if shape else "64, 128"
                lines.append(f'{indent}w_{config.id} = np.random.randn({shape[-1] if shape else 64}, 256).astype(np.{dtype})')
                lines.append(f'{indent}out_{config.id} = nn.linear({input_var}, w_{config.id})')
            else:
                lines.append(f'{indent}w_{config.id} = np.random.randn({input_var}.shape[-1], 256).astype(np.{dtype})')
                lines.append(f'{indent}out_{config.id} = nn.linear({input_var}, w_{config.id})')
        elif op_name == "matmul":
            if len(depends) >= 2:
                keys = list(depends.keys())
                lines.append(f'{indent}out_{config.id} = nn.matmul({keys[0]}_{config.id}, {keys[1]}_{config.id})')
            else:
                if not depends:
                    shape_str = ", ".join(str(d) for d in shape) if shape else "64, 64"
                    lines.append(f'{indent}b_{config.id} = np.random.randn({shape[-1] if shape else 64}, {shape[-1] if shape else 64}).astype(np.{dtype})')
                else:
                    lines.append(f'{indent}b_{config.id} = np.random.randn({input_var}.shape[-1], {input_var}.shape[-1]).astype(np.{dtype})')
                lines.append(f'{indent}out_{config.id} = nn.matmul({input_var}, b_{config.id})')
        elif op_name == "relu":
            lines.append(f'{indent}out_{config.id} = nn.relu({input_var})')
        elif op_name == "softmax":
            lines.append(f'{indent}out_{config.id} = nn.softmax({input_var})')
        elif op_name == "attention":
            if "q" in depends and "k" in depends and "v" in depends:
                lines.append(f'{indent}out_{config.id} = nn.attention(q_{config.id}, k_{config.id}, v_{config.id})')
            else:
                lines.append(f'{indent}# attention 需要 q, k, v 三个输入')
                lines.append(f'{indent}out_{config.id} = {input_var}  # placeholder')
        elif op_name == "add":
            if len(depends) >= 2:
                keys = list(depends.keys())
                lines.append(f'{indent}out_{config.id} = nn.add({keys[0]}_{config.id}, {keys[1]}_{config.id})')
            else:
                lines.append(f'{indent}out_{config.id} = nn.add({input_var}, {input_var})')
        elif op_name == "mul":
            if len(depends) >= 2:
                keys = list(depends.keys())
                lines.append(f'{indent}out_{config.id} = nn.mul({keys[0]}_{config.id}, {keys[1]}_{config.id})')
            else:
                lines.append(f'{indent}out_{config.id} = nn.mul({input_var}, {input_var})')
        else:
            # 默认：尝试调用 nn 模块的同名函数
            lines.append(f'{indent}# 未知算子: {op_name}')
            lines.append(f'{indent}out_{config.id} = {input_var}  # placeholder')

        # 保存输出
        lines.append(f'{indent}outputs[{config.id}] = out_{config.id}')
        lines.append('')

    # 生成导出代码
    lines.extend([
        '    # 导出结果',
        '    dump("./workspace")',
        '    csv_path = gen_csv("./workspace", "generated")',
        '    print(f"生成 compare 配置: {csv_path}")',
        '    return outputs',
        '',
        '',
        'if __name__ == "__main__":',
        '    run()',
        '',
    ])

    code = "\n".join(lines)

    # 写文件
    if output_py:
        output_path = Path(output_py)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(code, encoding="utf-8")
        logger.info(f"生成 Python 代码: {output_py}")

    return code
